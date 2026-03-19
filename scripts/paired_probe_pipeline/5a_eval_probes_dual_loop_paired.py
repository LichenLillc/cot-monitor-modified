"""
Script: 5a_eval_probes_dual_loop_paired.py (Ultimate Cache & MultiIndex Layout Version)

Description:
    Fully automated Matrix Evaluation for Probes.
    - [NEW] Directory Isolation: Separates human-readable Excel reports and joblib caches into distinct folders.
    - [NEW] Checkpoint Resume: Saves raw eval data as .joblib. Skips heavy CPU eval if cache exists.
    - [REFACTORED] Layout: Generates one independent Excel file per (Train_Model, Test_Model).
    - [REFACTORED] MultiIndex: Flawless double-header layout with thin borders for L1 group splits.
    - [OPTIMIZED] ID-TEST: Dynamically parses Markdown tables from 3a training summary.
    - Fast In-Memory Caching & Zero-Copy ProcessPool included.
"""

import os
# 防止多进程 CPU 过载
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["VECLIB_MAXIMUM_THREADS"] = "1"
os.environ["NUMEXPR_NUM_THREADS"] = "1"

import sys
import re
import json
import argparse
import pathlib
import collections
import joblib
import numpy as np
import pandas as pd
import torch
import torch.nn as nn
from tqdm import tqdm
from loguru import logger
import openpyxl 
from openpyxl.styles import Alignment, Border, Side
from sklearn.metrics import accuracy_score, f1_score, roc_auc_score, average_precision_score
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed

# ==========================================
# 0. Global Memory Cache (For Datasets)
# ==========================================
GLOBAL_CACHE = {}

# ==========================================
# 1. MLP Definitions
# ==========================================
class MLP_V1_Shallow(nn.Module):
    def __init__(self, input_size, dropout_rate=0.3):
        super().__init__()
        self.fc1 = nn.Linear(input_size, 128)
        self.relu = nn.ReLU()
        self.drop = nn.Dropout(dropout_rate)
        self.head = nn.Linear(128, 1)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        x = self.drop(self.relu(self.fc1(x)))
        return self.sigmoid(self.head(x))

class MLP_V2_Medium(nn.Module):
    def __init__(self, input_size, dropout_rate=0.3):
        super().__init__()
        self.fc1 = nn.Linear(input_size, 256)
        self.ln1 = nn.LayerNorm(256)
        self.relu = nn.ReLU()
        self.drop = nn.Dropout(dropout_rate)
        self.fc2 = nn.Linear(256, 128)
        self.ln2 = nn.LayerNorm(128)
        self.head = nn.Linear(128, 1)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        x = self.drop(self.relu(self.ln1(self.fc1(x))))
        x = self.drop(self.relu(self.ln2(self.fc2(x))))
        return self.sigmoid(self.head(x))

class MLP_V3_Deep(nn.Module):
    def __init__(self, input_size, dropout_rate=0.3):
        super().__init__()
        self.fc1 = nn.Linear(input_size, 512)
        self.ln1 = nn.LayerNorm(512)
        self.fc2 = nn.Linear(512, 256)
        self.ln2 = nn.LayerNorm(256)
        self.fc3 = nn.Linear(256, 128)
        self.ln3 = nn.LayerNorm(128)
        self.relu = nn.ReLU()
        self.drop = nn.Dropout(dropout_rate)
        self.head = nn.Linear(128, 1)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        x = self.drop(self.relu(self.ln1(self.fc1(x))))
        x = self.drop(self.relu(self.ln2(self.fc2(x))))
        x = self.drop(self.relu(self.ln3(self.fc3(x))))
        return self.sigmoid(self.head(x))

# ==========================================
# 2. Data Loading & Metrics
# ==========================================
def load_activations(input_folder):
    input_folder = pathlib.Path(input_folder)
    activations_path = input_folder / "activations"
    labels_path = input_folder / "labels"
    
    if not activations_path.exists() or not labels_path.exists():
        return None, None

    xs, ys = [], []
    pt_files = list(activations_path.glob("*.pt"))
    if not pt_files: return None, None

    for activation_file in pt_files: 
        parts = activation_file.stem.rsplit("_", 1)
        if len(parts) != 2: continue
        prompt_id, cot_id = parts
        
        label_file = labels_path / prompt_id / f"{prompt_id}_{cot_id}_labeled.json"
        if not label_file.exists(): continue
            
        try:
            act = torch.load(activation_file, map_location='cpu')
            if act.dim() > 1: act = act.view(-1)
            xs.append(act.numpy())
            with open(label_file, 'r') as f:
                score = json.load(f)["safety_label"]["score"]
                ys.append(1 if score >= 0.5 else 0)
        except Exception:
            continue
            
    if len(xs) == 0: return None, None
    return np.array(xs), np.array(ys)

def preload_single_dataset(d_path):
    xs, ys = load_activations(d_path)
    if xs is not None and ys is not None:
        return str(d_path), xs, ys
    return str(d_path), None, None

def calculate_metrics(y_true, y_pred, y_probs):
    metrics = {"accuracy": float(accuracy_score(y_true, y_pred)) * 100}
    if len(np.unique(y_true)) > 1:
        metrics["f1"] = float(f1_score(y_true, y_pred)) * 100
        metrics["auc_roc"] = float(roc_auc_score(y_true, y_probs)) * 100
        metrics["pr_auc"] = float(average_precision_score(y_true, y_probs)) * 100
    return metrics

def get_available_seeds(model_folder_path):
    model_dir = model_folder_path / "saved_models"
    if not model_dir.exists(): return []
    seeds = sorted(list(set(
        int(re.search(r"seed(\d+)", p.name).group(1)) 
        for p in model_dir.glob("global_scaler_seed*.joblib") if re.search(r"seed(\d+)", p.name)
    )))
    return seeds

def get_id_test_accuracy(model_folder):
    """
    Dynamically parses the Markdown table in training_summary.txt
    to extract accurately formatted metric strings like "95.1 ± 1.2"
    """
    summary_path = model_folder / "training_summary.txt"
    results = {"raw": {}, "pca": {}}
    for m in ["logreg", "mlp_v1", "mlp_v2", "mlp_v3"]:
        results["raw"][m], results["pca"][m] = "N/A", "N/A"
        
    if not summary_path.exists(): return results
        
    try:
        content = summary_path.read_text(encoding='utf-8')
        lines = content.splitlines()
        acc_idx = -1
        
        for line in lines:
            if "Model" in line and "accuracy" in line.lower():
                headers = [h.strip().lower() for h in line.split('|')]
                if "accuracy" in headers: acc_idx = headers.index("accuracy")
                break
                
        if acc_idx != -1:
            for line in lines:
                if line.strip().startswith("`"): 
                    line_lower = line.lower()
                    for mode in ["raw", "pca"]:
                        for m in ["logreg", "mlp_v1", "mlp_v2", "mlp_v3"]:
                            key = f"{m}_{mode}"
                            if f"`{key}" in line_lower:
                                parts = [p.strip() for p in line.split('|')]
                                if len(parts) > acc_idx:
                                    results[mode][m] = parts[acc_idx]
        return results
    except Exception as e:
        logger.warning(f"Error parsing ID-TEST for {model_folder.name}: {e}")
        return results

# ==========================================
# 3. Core Task: Evaluate ONE Model-Seed
# ==========================================
def evaluate_model_seed_all_datasets(m_path, row_key, dataset_info_list, seed):
    model_dir = m_path / "saved_models"
    scaler_path = model_dir / f"global_scaler_seed{seed}.joblib"
    if not scaler_path.exists(): return None
    global_scaler = joblib.load(scaler_path)
    
    pca_path = model_dir / f"pca_model_seed{seed}.joblib"
    pca_model = joblib.load(pca_path) if pca_path.exists() else None
    
    device = torch.device("cpu")
    model_types = ["logreg", "mlp_v1", "mlp_v2", "mlp_v3"]
    loaded_models = {"raw": {}, "pca": {}}
    
    for mode in ["raw", "pca"]:
        for m_type in model_types:
            if m_type == "logreg":
                p = model_dir / f"logreg_{mode}_seed{seed}.joblib"
                if p.exists(): loaded_models[mode][m_type] = joblib.load(p)
            else:
                p = model_dir / f"{m_type}_{mode}_seed{seed}.pth"
                scl_p = model_dir / f"{m_type}_scaler_{mode}_seed{seed}.joblib"
                if p.exists() and scl_p.exists():
                    loaded_models[mode][m_type] = (torch.load(p, map_location=device), joblib.load(scl_p))

    results = {}
    for d_path_str, col_key in dataset_info_list:
        if d_path_str not in GLOBAL_CACHE: continue
        X_raw, y_test = GLOBAL_CACHE[d_path_str]
        
        try: X_stage1 = global_scaler.transform(X_raw)
        except ValueError: continue # 柔性盾牌：维度不匹配跳过

        X_pca = pca_model.transform(X_stage1) if pca_model else None
        res_d = {"raw": {}, "pca": {}}
        
        for mode in ["raw", "pca"]:
            X_curr = X_stage1 if mode == "raw" else X_pca
            if X_curr is None: continue
            
            for m_type in model_types:
                if m_type not in loaded_models[mode]: continue
                
                try:
                    if m_type == "logreg":
                        mod = loaded_models[mode][m_type]
                        y_probs = mod.predict_proba(X_curr)[:, 1]
                        y_pred = mod.predict(X_curr)
                        res_d[mode][m_type] = calculate_metrics(y_test, y_pred, y_probs)
                    else:
                        state_dict, mlp_scaler = loaded_models[mode][m_type]
                        X_in = mlp_scaler.transform(X_curr)
                        in_dim = X_in.shape[1]
                        
                        if m_type == "mlp_v1": mlp = MLP_V1_Shallow(in_dim)
                        elif m_type == "mlp_v2": mlp = MLP_V2_Medium(in_dim)
                        elif m_type == "mlp_v3": mlp = MLP_V3_Deep(in_dim)
                        
                        mlp.load_state_dict(state_dict)
                        mlp.eval()
                        with torch.no_grad():
                            probs = mlp(torch.tensor(X_in, dtype=torch.float32)).squeeze().numpy()
                            if probs.ndim == 0: probs = np.expand_dims(probs, 0)
                            y_pred = (probs >= 0.5).astype(int)
                        res_d[mode][m_type] = calculate_metrics(y_test, y_pred, probs)
                except Exception: continue
        results[col_key] = res_d
    return row_key, seed, results

# ==========================================
# 4. Helper: Name Parsing
# ==========================================
def parse_folder_name(folder_name, suffix_to_remove):
    name = folder_name
    if suffix_to_remove and name.endswith(f"_{suffix_to_remove}"):
        name = name[:-(len(suffix_to_remove) + 1)] 
    name = re.sub(r'_?test_?', '_', name, flags=re.IGNORECASE)
    
    l1 = "Others"
    if "7b_pfc_think-ins" in name: l1 = "qwen7b-pfc"
    elif "ds-coder-ckpt280" in name: l1 = "ds-coder"
    elif "qwen_scratch-ckpt61" in name: l1 = "qwen-unittest"
    elif "qwen_exit-ckpt68" in name: l1 = "qwen-exit"
    elif "wild" in name: l1 = "qwen-exit-scratch"

    keywords = ["7b_pfc_think-ins_cot", "7b_pfc_think-ins", "ds-coder-ckpt280", "qwen_exit-ckpt68", "qwen_scratch-ckpt61", "wild_dup4"]
    l2_candidate = name
    for k in keywords: l2_candidate = l2_candidate.replace(k, "")
    l2 = re.sub(r'_+', '_', l2_candidate).strip('_')
    if not l2: l2 = "base"
    return l1, l2

def get_train_row_weight(name):
    name_lower = name.lower()
    is_unpaired = 1 if 'unpaired' in name_lower else 0
    clean_name = re.sub(r'ckpt\d+', '', name_lower)
    clean_name = re.sub(r'7b_pfc|ds-coder|qwen_scratch|qwen_exit|wild_dup4', '', clean_name)
    parts = re.findall(r'[a-z]+(\d+)', clean_name)
    num_types = -len(parts)
    total_data = -sum([int(x) for x in parts])
    return (is_unpaired, num_types, total_data, name_lower)

def get_test_col_weight(name):
    name_lower = name.lower()
    if 'ln' in name_lower: return 1
    if 'tn' in name_lower: return 2
    if 'uh' in name_lower: return 3
    if 'eh' in name_lower: return 4
    if 'sh' in name_lower: return 5
    if 'hh' in name_lower: return 6
    if 'mh' in name_lower: return 7
    return 8

# ==========================================
# 5. Excel Generation (MultiIndex + Thin Borders)
# ==========================================
def generate_pair_excel_report(mp_name, dp_name, cfg, modality_name, reports_dir):
    excel_path = reports_dir / f"Train_{mp_name}_on_Test_{dp_name}_{modality_name}.xlsx"
    
    L1_ORDER = ["qwen-unittest", "qwen-exit", "qwen-exit-scratch", "ds-coder", "qwen7b-pfc"]
    
    # [核心修复]：在这里补上 L1_ORDER 的排序权重！
    sorted_rows = sorted(cfg["model_rows"], key=lambda x: (L1_ORDER.index(x[0]) if x[0] in L1_ORDER else 99, get_train_row_weight(x[1])))
    
    sorted_cols = sorted(cfg["dataset_cols"], key=lambda x: (L1_ORDER.index(x[0]) if x[0] in L1_ORDER else 99, get_test_col_weight(x[1]), x[1]))

    row_tuples = [(r[0], r[1]) for r in sorted_rows]
    col_tuples = [('ID-TEST', '')] + [(c[0], c[1]) for c in sorted_cols]
    
    row_index = pd.MultiIndex.from_tuples(row_tuples, names=['Train Type', 'Train Detail'])
    col_index = pd.MultiIndex.from_tuples(col_tuples, names=['Test Type', 'Test Detail'])

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        for m_type in ["logreg", "mlp_v1", "mlp_v2", "mlp_v3"]:
            sheet_name = m_type.upper()
            row_offset = 0

            for mode in ["raw", "pca"]:
                df = pd.DataFrame(index=row_index, columns=col_index)

                for r in sorted_rows:
                    r_key = (r[0], r[1])
                    df.loc[r_key, ('ID-TEST', '')] = cfg["report_data"][mode][m_type][r_key].get("ID-TEST", "N/A")
                    for c in sorted_cols:
                        c_key = (c[0], c[1])
                        df.loc[r_key, c_key] = cfg["report_data"][mode][m_type][r_key].get(c_key, "N/A")

                df.to_excel(writer, sheet_name=sheet_name, startrow=row_offset + 1)
                worksheet = writer.sheets[sheet_name]
                
                title = f"[ {mode.upper()} FEATURES ]"
                worksheet.cell(row=row_offset + 1, column=1, value=title).font = openpyxl.styles.Font(bold=True, size=14)
                
                row_offset += len(df) + 5

    workbook = openpyxl.load_workbook(excel_path)
    thin_side = Side(style='thin')
    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        L1_col_starts = []
        current_L1 = None
        for i, col in enumerate(col_tuples):
            l1 = col[0]
            if l1 != current_L1:
                L1_col_starts.append(i + 3) # Offset: 2 for Row Headers + 1 (1-based index)
                current_L1 = l1
                
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                if cell.column in L1_col_starts:
                    existing = cell.border
                    cell.border = Border(left=thin_side, right=existing.right, top=existing.top, bottom=existing.bottom)
                    
    workbook.save(excel_path)

# ==========================================
# 6. Global Setup & Execution
# ==========================================
def process_modality(mgp_path, dgp_path, modality_name, args):
    results_root = pathlib.Path(args.results_root)
    
    # 隔离 reports 目录和 cache 目录
    reports_dir = results_root / "reports" / modality_name
    cache_dir = results_root / "cache" / modality_name
    reports_dir.mkdir(parents=True, exist_ok=True)
    cache_dir.mkdir(parents=True, exist_ok=True)
    
    mps = [d for d in mgp_path.iterdir() if d.is_dir()]
    dps = [d for d in dgp_path.iterdir() if d.is_dir()]
    L1_ORDER = ["qwen-unittest", "qwen-exit", "qwen-exit-scratch", "ds-coder", "qwen7b-pfc"]
    
    global_tasks = []
    matrix_configs = {}
    datasets_to_preload = set()
    
    # 1. Checkpoint Scanning & Task Assembly
    for mp in mps:
        for dp in dps:
            models = [d for d in mp.iterdir() if d.is_dir() and (d / "saved_models").exists()]
            datasets = [d for d in dp.iterdir() if d.is_dir() and (d / "activations").exists() and "test" in d.name.lower()]
            if not models or not datasets: continue

            model_rows, dataset_cols = [], [] 
            for m in models:
                l1, l2 = parse_folder_name(m.name, mp.name)
                if l1 != "Others" and l1 in L1_ORDER:
                    model_rows.append((l1, l2, m, get_id_test_accuracy(m)))
            
            for d in datasets:
                l1, l2 = parse_folder_name(d.name, dp.name)
                if l1 != "Others" and l1 in L1_ORDER:
                    dataset_cols.append((l1, l2, str(d)))

            if not model_rows or not dataset_cols: continue
            
            cache_file = cache_dir / f"Cache_{mp.name}_on_{dp.name}.joblib"
            
            cfg = {
                "model_rows": model_rows,
                "dataset_cols": dataset_cols,
                "is_cached": False,
                "agg_metrics": {"raw": collections.defaultdict(lambda: collections.defaultdict(lambda: collections.defaultdict(list))),
                                "pca": collections.defaultdict(lambda: collections.defaultdict(lambda: collections.defaultdict(list)))},
                "report_data": {"raw": collections.defaultdict(dict), "pca": collections.defaultdict(dict)}
            }
            
            if cache_file.exists():
                logger.info(f"[{modality_name}] Loading CACHED results for {mp.name} on {dp.name}")
                cfg["report_data"] = joblib.load(cache_file)
                cfg["is_cached"] = True
            else:
                dataset_info_list = [(d_str, (l1, l2)) for l1, l2, d_str in dataset_cols]
                for ml1, ml2, m_path, id_vals in model_rows:
                    row_key = (ml1, ml2)
                    for mode in ["raw", "pca"]:
                        for m_type in ["logreg", "mlp_v1", "mlp_v2", "mlp_v3"]:
                            cfg["report_data"][mode][m_type][row_key] = {"ID-TEST": id_vals[mode][m_type]}
                            for _, col_key in dataset_info_list:
                                cfg["report_data"][mode][m_type][row_key][col_key] = "N/A"

                    seeds = [args.seed] if args.seed is not None else get_available_seeds(m_path)
                    for seed in seeds:
                        global_tasks.append({
                            "mp_name": mp.name, "dp_name": dp.name,
                            "m_path": m_path, "row_key": row_key, 
                            "dataset_info_list": dataset_info_list, "seed": seed
                        })
                # Only preload datasets that are actually needed for computation
                for _, _, d_str in dataset_cols:
                    datasets_to_preload.add(pathlib.Path(d_str))
                    
            matrix_configs[(mp.name, dp.name)] = cfg

    # 2. Lazy Preloading (Only preload if there are uncached tasks)
    if datasets_to_preload:
        logger.info(f"[{modality_name}] Fast Preloading {len(datasets_to_preload)} unique datasets into Global RAM...")
        with ThreadPoolExecutor(max_workers=32) as io_pool:
            futures = {io_pool.submit(preload_single_dataset, d): d for d in datasets_to_preload}
            for future in tqdm(as_completed(futures), total=len(datasets_to_preload), desc=f"{modality_name} IO Read"):
                d_str, xs, ys = future.result()
                if xs is not None: GLOBAL_CACHE[d_str] = (xs, ys)
    else:
        logger.info(f"[{modality_name}] All pairs cached! Skipping IO Preloading.")

    # 3. ProcessPool Evaluation (For uncached pairs)
    if global_tasks:
        logger.info(f"[{modality_name}] Launching {len(global_tasks)} Optimized Tasks via ProcessPool ({args.workers} workers)...")
        with ProcessPoolExecutor(max_workers=args.workers) as executor:
            future_to_task = {
                executor.submit(evaluate_model_seed_all_datasets, t["m_path"], t["row_key"], t["dataset_info_list"], t["seed"]): t
                for t in global_tasks
            }
            for future in tqdm(as_completed(future_to_task), total=len(global_tasks), desc=f"{modality_name} Fast Eval"):
                task = future_to_task[future]
                try:
                    res = future.result()
                    if res:
                        row_key, seed, results_dict = res
                        cfg = matrix_configs[(task["mp_name"], task["dp_name"])]
                        for col_key, modes_dict in results_dict.items():
                            for mode, m_types_dict in modes_dict.items():
                                for m_type, metrics in m_types_dict.items():
                                    if "accuracy" in metrics:
                                        cfg["agg_metrics"][mode][m_type][row_key][col_key].append(metrics["accuracy"])
                except Exception as e:
                    logger.error(f"Task failed: {e}")

    # 4. Save Cache & Generate Excel
    for (mp_name, dp_name), cfg in matrix_configs.items():
        if not cfg["is_cached"]:
            for mode in ["raw", "pca"]:
                for m_type in ["logreg", "mlp_v1", "mlp_v2", "mlp_v3"]:
                    for row_key, col_dict in cfg["agg_metrics"][mode][m_type].items():
                        for col_key, acc_list in col_dict.items():
                            if acc_list:
                                cfg["report_data"][mode][m_type][row_key][col_key] = f"{np.mean(acc_list):.2f} ± {np.std(acc_list):.2f}"
            
            # Save Checkpoint File into Cache Folder
            cache_file = cache_dir / f"Cache_{mp_name}_on_{dp_name}.joblib"
            joblib.dump(cfg["report_data"], cache_file)
            logger.info(f"Saved Checkpoint: {cache_file.name}")
        
        # Always generate fresh Excel from memory (whether newly computed or cached)
        generate_pair_excel_report(mp_name, dp_name, cfg, modality_name, reports_dir)
        
    logger.success(f"[{modality_name}] All Pairwise Reports Generated Successfully!")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mgp_text", type=str, default=None, help="Model grandparent folder for TEXT")
    parser.add_argument("--dgp_text", type=str, default=None, help="Dataset grandparent folder for TEXT")
    parser.add_argument("--mgp_eos", type=str, default=None, help="Model grandparent folder for EOS")
    parser.add_argument("--dgp_eos", type=str, default=None, help="Dataset grandparent folder for EOS")
    parser.add_argument("--seed", type=int, default=None)
    parser.add_argument("--results_root", type=str, default="/data/lichenli/cot-monitor-modified/main_table3_paired/exp_0314/5a_results_0314/")
    parser.add_argument("--workers", type=int, default=50, help="Number of parallel workers")
    args = parser.parse_args()
    
    if args.mgp_text and args.dgp_text:
        process_modality(pathlib.Path(args.mgp_text), pathlib.Path(args.dgp_text), "TEXT", args)
        
    if args.mgp_eos and args.dgp_eos:
        process_modality(pathlib.Path(args.mgp_eos), pathlib.Path(args.dgp_eos), "EOS", args)

if __name__ == "__main__":
    main()